import os
import pandas as pd
from dados import extrair_dados
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter


# Caminho da pasta com os arquivos XML
pasta_xml = "./orcamentos"
arquivos_xml = [arquivo for arquivo in os.listdir(pasta_xml) if arquivo.endswith(".xml")]

# Listas para armazenar dados
todos_os_dados = []
todos_os_servicos = []

# Extrai os dados de cada XML
for arquivo in arquivos_xml:
    caminho_arquivo = os.path.join(pasta_xml, arquivo)
    dados = extrair_dados(caminho_arquivo)
    dados["arquivo"] = arquivo

    servicos_terceirizados = []
    total_terceirizados = 0.0

    for item in dados.get("Serviços tercerizados: \n", "").split("\n\n"):
        if item.strip():
            nome_servico, preco = item.split("\n")
            preco = preco.replace("R$ ", "").replace(" ", "").replace(",", ".")
            preco_float = float(preco)
            servicos_terceirizados.append({"nome": nome_servico, "preco": preco_float, "arquivo": arquivo})
            total_terceirizados += preco_float

    dados["total_terceirizados"] = total_terceirizados

    for servico in servicos_terceirizados:
        todos_os_servicos.append(servico)

    todos_os_dados.append(dados)

# Cria DataFrame e renomeia colunas
df_dados = pd.DataFrame(todos_os_dados)
df_dados.rename(columns={
    "seguradora:": "Seguradora",
    "marca:": "Marca",
    "veiculo:": "Veículo",
    "placa:": "Placa",
    "pecas oficina:": "Peças Oficina ",
    "funilaria:": "Funilaria ",
    "pintura:": "Pintura ",
    "mecanica:": "Mecânica ",
    "montagem desmontagem:": "MONT/DESM",
    "total_terceirizados": "Serviços ",
    "Total mão de obra:": "Mão de Obra Total ",
    "franquia:": "Franquia ",
    "total_liquido_geral:": "Total Geral "
}, inplace=True)

# Reorganiza a ordem das colunas
colunas_ordenadas = [
    "Seguradora", "Marca", "Veículo", "Placa",
    "Peças Oficina ", "Funilaria ", "Pintura ",
    "Mecânica ", "MONT/DESM",
    "Serviços ", "Mão de Obra Total ",
    "Franquia ", "Total Geral ", "arquivo"
]
df_dados = df_dados[colunas_ordenadas]

# DataFrame de serviços
df_servicos = pd.DataFrame(todos_os_servicos)

# Salvar a planilha
pasta_saida = "./planilhas"
os.makedirs(pasta_saida, exist_ok=True)
caminho_planilha = os.path.join(pasta_saida, "PlanilhaFechamento.xlsx")
# Caminho da planilha base com gráficos
planilha_base = "planilha_modelo.xlsx"
wb = load_workbook(planilha_base)

# Atualiza aba da página 1 com novos dados
aba_dados = wb["Controle carros"]  # Nome da aba com dados a serem substituídos

# Limpa os dados antigos (mantém cabeçalho)
for row in aba_dados.iter_rows(min_row=2, max_row=aba_dados.max_row):
    for cell in row:
        cell.value = None

# Insere os novos dados no local
from openpyxl.utils.dataframe import dataframe_to_rows
for r_idx, row in enumerate(dataframe_to_rows(df_dados, index=False, header=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        cell = aba_dados.cell(row=r_idx, column=c_idx, value=value)
        if isinstance(value, float):
            cell.number_format = 'R$ #,##0.00'
        cell.alignment = Alignment(horizontal='center')

# Estilo do cabeçalho
for cell in aba_dados[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Ajusta largura automática
from openpyxl.utils import get_column_letter
for col in aba_dados.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    aba_dados.column_dimensions[col_letter].width = max_length + 2

# Atualiza segunda aba com serviços terceirizados
if "Serviços Terceirizados" not in wb.sheetnames:
    wb.create_sheet("Serviços Terceirizados")

aba_servicos = wb["Serviços Terceirizados"]

# Limpa serviços antigos
for row in aba_servicos.iter_rows(min_row=2, max_row=aba_servicos.max_row):
    for cell in row:
        cell.value = None

# Escreve serviços detalhados
for r_idx, row in enumerate(dataframe_to_rows(df_servicos, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        aba_servicos.cell(row=r_idx, column=c_idx, value=value)

# Salva a nova planilha
saida_final = os.path.join(pasta_saida, "PlanilhaFechamento.xlsx")
wb.save(saida_final)
print(f"Planilha atualizada com sucesso em: {saida_final}")
